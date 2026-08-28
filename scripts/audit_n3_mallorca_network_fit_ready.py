from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path

from openpyxl import load_workbook

DOI = "10.5061/dryad.hqbzkh1bm"
EXPECTED_FILES = {"Dryad.xlsx", "Dryad_notes.docx"}
NETWORK_SHEETS = {"2016": "Sheet1_Network2016", "2017": "Sheet2_Network2017"}
SEM_SHEET = "Sheet 3_SEMvariables"
REQUIRED_SEM_COLUMNS = {"Year", "Species", "DPD", "FloralUnitSize", "SeedsFlowerRounded"}
MIN_DISTINCT_SPECIES = 15
MIN_SPECIES_YEAR_PAIRS = 25
MIN_PAIRS_PER_YEAR = 10


def _req(url: str, json_only: bool = False) -> urllib.request.Request:
    return urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
            "Accept": "application/json" if json_only else "application/zip,application/octet-stream,*/*;q=0.8",
        },
    )


def _get_json(url: str) -> dict:
    with urllib.request.urlopen(_req(url, json_only=True), timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def _norm_species(value: object) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _year(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def _download_locked_workbook() -> tuple[bytes, dict]:
    encoded = urllib.parse.quote(f"doi:{DOI}", safe="")
    versions = _get_json(f"https://datadryad.org/api/v2/datasets/{encoded}/versions").get("_embedded", {}).get("stash:versions", [])
    if not versions:
        raise RuntimeError("no Dryad versions")
    latest = versions[-1]
    files_href = latest["_links"]["stash:files"]["href"]
    download_href = latest["_links"]["stash:download"]["href"]
    files_url = "https://datadryad.org" + files_href if files_href.startswith("/") else files_href
    items = _get_json(files_url).get("_embedded", {}).get("stash:files", [])
    manifest = {
        str(item["path"]): {"size": int(item["size"]), "digest": str(item["digest"])}
        for item in items
    }
    missing = EXPECTED_FILES - set(manifest)
    if missing:
        raise RuntimeError(f"missing locked archive files: {sorted(missing)}")
    download_url = "https://datadryad.org" + download_href if download_href.startswith("/") else download_href
    with urllib.request.urlopen(_req(download_url), timeout=120) as response:
        bundle = response.read()
        final_url = response.geturl()
    with zipfile.ZipFile(io.BytesIO(bundle)) as zf:
        by_name = {Path(name).name: name for name in zf.namelist()}
        workbook = zf.read(by_name["Dryad.xlsx"])
        notes = zf.read(by_name["Dryad_notes.docx"])
    verification = {}
    for name, data in (("Dryad.xlsx", workbook), ("Dryad_notes.docx", notes)):
        sha = hashlib.sha256(data).hexdigest()
        expected = manifest[name]
        accepted = len(data) == expected["size"] and sha == expected["digest"]
        if not accepted:
            raise RuntimeError(f"locked identity mismatch for {name}")
        verification[name] = {"bytes": len(data), "sha256": sha, "accepted": True}
    return workbook, {
        "version": latest["_links"]["self"]["href"],
        "download_final_url": final_url,
        "file_verification": verification,
    }


def audit() -> dict:
    payload, provenance = _download_locked_workbook()
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    required_sheets = set(NETWORK_SHEETS.values()) | {SEM_SHEET}
    missing_sheets = sorted(required_sheets - set(wb.sheetnames))
    if missing_sheets:
        raise RuntimeError(f"missing required sheets: {missing_sheets}")

    network_species: dict[str, set[str]] = {}
    network_schema = {}
    for year, sheet_name in NETWORK_SHEETS.items():
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        first_header = str(header[0] or "").strip()
        pollinator_columns = sum(1 for value in header[1:] if str(value or "").strip())
        plants = set()
        rows_with_any_link_cell = 0
        for row in rows:
            species = _norm_species(row[0] if row else None)
            if not species:
                continue
            plants.add(species)
            # Only missingness is inspected; numeric interaction values are not surfaced or associated with F.
            if any(value is not None for value in row[1:]):
                rows_with_any_link_cell += 1
        network_species[year] = plants
        network_schema[year] = {
            "sheet": sheet_name,
            "first_header": first_header,
            "plant_row_count": len(plants),
            "pollinator_column_count": pollinator_columns,
            "plant_rows_with_any_link_cell": rows_with_any_link_cell,
        }

    ws = wb[SEM_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    missing_columns = sorted(REQUIRED_SEM_COLUMNS - set(header))
    if missing_columns:
        raise RuntimeError(f"missing SEM columns: {missing_columns}")
    index = {name: header.index(name) for name in REQUIRED_SEM_COLUMNS}

    sem_pairs = []
    exclusions = {
        "missing_or_unknown_year": [],
        "missing_species": [],
        "species_absent_from_year_network": [],
        "missing_DPD": [],
        "missing_FloralUnitSize": [],
        "missing_SeedsFlowerRounded": [],
    }
    for row in rows:
        year = _year(row[index["Year"]])
        species = _norm_species(row[index["Species"]])
        pair = f"{year}|{species}" if year and species else ""
        if year not in NETWORK_SHEETS:
            if pair:
                exclusions["missing_or_unknown_year"].append(pair)
            continue
        if not species:
            exclusions["missing_species"].append(year)
            continue
        if species not in network_species[year]:
            exclusions["species_absent_from_year_network"].append(pair)
            continue
        if row[index["DPD"]] is None:
            exclusions["missing_DPD"].append(pair)
            continue
        if row[index["FloralUnitSize"]] is None:
            exclusions["missing_FloralUnitSize"].append(pair)
            continue
        if row[index["SeedsFlowerRounded"]] is None:
            exclusions["missing_SeedsFlowerRounded"].append(pair)
            continue
        sem_pairs.append((year, species))

    pair_counts = {year: sum(1 for y, _ in sem_pairs if y == year) for year in NETWORK_SHEETS}
    species = sorted({species for _, species in sem_pairs})
    eligible = (
        len(sem_pairs) >= MIN_SPECIES_YEAR_PAIRS
        and len(species) >= MIN_DISTINCT_SPECIES
        and all(pair_counts[year] >= MIN_PAIRS_PER_YEAR for year in NETWORK_SHEETS)
    )
    return {
        "analysis": "N3_Mallorca_network_fit_ready_stage_A",
        "doi": DOI,
        "decision": "fit_ready_for_prospective_B1_preregistration" if eligible else "stage_A_not_fit_ready",
        "thresholds_declared_before_outcome_modeling": {
            "min_distinct_species": MIN_DISTINCT_SPECIES,
            "min_species_year_pairs": MIN_SPECIES_YEAR_PAIRS,
            "min_pairs_per_year": MIN_PAIRS_PER_YEAR,
        },
        "network_schema": network_schema,
        "sem_schema": {
            "sheet": SEM_SHEET,
            "required_columns_present": True,
            "eligible_species_year_pair_count": len(sem_pairs),
            "eligible_distinct_species_count": len(species),
            "eligible_pair_counts_by_year": pair_counts,
            "eligible_species_identifiers": species,
            "exclusions_by_reason": {key: sorted(values) for key, values in exclusions.items()},
        },
        "provenance": provenance,
        "response_firewall": "Stage A inspected archive identity, headers, units implied by source documentation, process-row presence, response/predictor missingness, and species-year join structure only. It did not compute direct-visitation row sums, inspect SeedsFlowerRounded values, calculate process-function associations, fit models, or rank candidate effects.",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="artifacts/empirical/n3_mallorca_network_fit_ready_stage_a.json")
    args = parser.parse_args()
    try:
        result = audit()
    except Exception as exc:
        result = {
            "analysis": "N3_Mallorca_network_fit_ready_stage_A",
            "doi": DOI,
            "decision": "stage_A_audit_error",
            "error": f"{type(exc).__name__}: {exc}",
            "response_firewall": "No process-function association or model was computed.",
        }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"decision": result["decision"], "sem_schema": result.get("sem_schema")}, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
