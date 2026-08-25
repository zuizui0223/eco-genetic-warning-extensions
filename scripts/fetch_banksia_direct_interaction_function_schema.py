from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import zipfile
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

import openpyxl
import xlrd

SOURCE_REPO = "stanwawrzyczek/Pollination-of-Banksia-catoglypta-Data"
SOURCE_COMMIT = "1ab685d62d005865935435bbd49cadba50080741"
SOURCE_PATH = "Banksia catoglypta pollination - DATA.zip"
SOURCE_BLOB_SHA = "91cc21eb4d967b702bd18f87f91be1b52cacb6a3"
PUBLICATION_DOI = "10.1093/botlinnean/boae024"
UA = "eco-genetic-warning-extensions/1.0"


def _raw_url() -> str:
    encoded = quote(SOURCE_PATH, safe="")
    return f"https://raw.githubusercontent.com/{SOURCE_REPO}/{SOURCE_COMMIT}/{encoded}"


def _download() -> tuple[str, bytes]:
    url = _raw_url()
    req = Request(url, headers={"User-Agent": UA, "Accept": "application/zip,application/octet-stream,*/*"})
    with urlopen(req, timeout=180) as response:
        payload = response.read()
    return url, payload


def _git_blob_sha(payload: bytes) -> str:
    prefix = f"blob {len(payload)}\0".encode("ascii")
    return hashlib.sha1(prefix + payload).hexdigest()


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise RuntimeError("could not decode table header")


def _table_header(raw: bytes, suffix: str) -> list[str]:
    text = _decode_text(raw)
    lines = text.splitlines()
    if not lines:
        return []
    first = lines[0]
    if suffix in {".tsv", ".tab"}:
        delimiter = "\t"
    elif suffix == ".txt":
        delimiter = "\t" if "\t" in first else ","
    else:
        delimiter = ","
    return [str(value) for value in next(csv.reader([first], delimiter=delimiter), [])]


def _xlsx_schema(raw: bytes) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    sheets: list[dict[str, object]] = []
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        sheets.append(
            {
                "sheet": ws.title,
                "rows_including_header": int(ws.max_row or 0),
                "columns_count": int(ws.max_column or 0),
                "columns": ["" if value is None else str(value) for value in first],
            }
        )
    wb.close()
    return sheets


def _xls_schema(raw: bytes) -> list[dict[str, object]]:
    wb = xlrd.open_workbook(file_contents=raw, on_demand=True)
    sheets: list[dict[str, object]] = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        sheets.append(
            {
                "sheet": name,
                "rows_including_header": int(ws.nrows),
                "columns_count": int(ws.ncols),
                "columns": [str(value) for value in (ws.row_values(0) if ws.nrows else [])],
            }
        )
    wb.release_resources()
    return sheets


def _member_schema(name: str, raw: bytes) -> dict[str, object]:
    suffix = Path(name).suffix.lower()
    row: dict[str, object] = {
        "member_name": name,
        "member_bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
    }
    if suffix in {".csv", ".tsv", ".tab", ".txt"}:
        row["columns"] = _table_header(raw, suffix)
    elif suffix == ".xlsx":
        row["workbook_schema"] = _xlsx_schema(raw)
    elif suffix == ".xls":
        row["workbook_schema"] = _xls_schema(raw)
    return row


def discover(manifest_path: Path) -> dict[str, object]:
    url, payload = _download()
    observed_blob = _git_blob_sha(payload)
    if observed_blob != SOURCE_BLOB_SHA:
        raise RuntimeError(
            f"fixed Git blob mismatch: expected={SOURCE_BLOB_SHA}, observed={observed_blob}"
        )
    if not zipfile.is_zipfile(io.BytesIO(payload)):
        raise RuntimeError("fixed Banksia source blob is not a ZIP archive")

    members: list[dict[str, object]] = []
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            raw = archive.read(info)
            members.append(_member_schema(info.filename, raw))

    result: dict[str, object] = {
        "status": "schema_only_discovery_complete",
        "publication_doi": PUBLICATION_DOI,
        "source_repository": SOURCE_REPO,
        "source_commit": SOURCE_COMMIT,
        "source_path": SOURCE_PATH,
        "source_git_blob_sha": SOURCE_BLOB_SHA,
        "observed_git_blob_sha": observed_blob,
        "raw_url": url,
        "zip_bytes": len(payload),
        "zip_sha256": hashlib.sha256(payload).hexdigest(),
        "members": members,
        "inspection_boundary": (
            "Only the fixed GitHub snapshot provenance, ZIP/member names/sizes/hashes, and first table/workbook header rows were inspected. "
            "No data-row value, visitation rate, visitor frequency, fruit/seed/seedling outcome, exclusion-treatment effect, coefficient, p-value, correlation or descriptive outcome statistic was read or computed."
        ),
        "next_gate": (
            "Using filenames and header labels only, map observation/site/plant/inflorescence/time keys, I_realised, access/exclusion state, and F_reproduction. "
            "Classify direct_IF_joint_state_identifiable, direct_IF_partial_state_identifiable, or direct_IF_not_identifiable before any outcome row is read."
        ),
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", default="artifacts/empirical/banksia_direct_interaction_function_schema.json")
    args = parser.parse_args()
    result = discover(Path(args.manifest))
    print(
        json.dumps(
            {
                "status": result["status"],
                "source_commit": result["source_commit"],
                "source_git_blob_sha": result["source_git_blob_sha"],
                "members": [row["member_name"] for row in result["members"]],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
