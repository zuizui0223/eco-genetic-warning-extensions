from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import quote, urljoin
from urllib.request import Request, urlopen

import openpyxl
import xlrd

DOI = "10.5061/dryad.f8539"
DRYAD_ROOT = "https://datadryad.org"
API_ROOT = f"{DRYAD_ROOT}/api/v2"
FILES = (
    ("pollinators_all.xls", "I_realised"),
    ("FruitSet.xlsx", "F_reproduction"),
    ("paternity.xlsx", "G_mating_C_pollen"),
    ("abortion_data_2011.xlsx", "R_offspring_cost"),
)
UA = "eco-genetic-warning-extensions/1.0"


def _request(url: str, *, accept: str) -> bytes:
    req = Request(
        url,
        headers={
            "User-Agent": UA,
            "Accept": accept,
            "Referer": "https://datadryad.org/",
            "X-API-Version": "2.1.0",
        },
    )
    with urlopen(req, timeout=180) as response:
        return response.read()


def _json(url: str) -> dict:
    return json.loads(_request(url, accept="application/json").decode("utf-8"))


def _absolute(href: str) -> str:
    return urljoin(DRYAD_ROOT, href)


def _href(obj: dict, relation: str) -> str | None:
    value = obj.get("_links", {}).get(relation)
    if isinstance(value, dict) and value.get("href"):
        return _absolute(str(value["href"]))
    return None


def _file_id(item: dict) -> int | None:
    if item.get("id") is not None:
        return int(item["id"])
    href = _href(item, "self")
    if href:
        match = re.search(r"/files/(\d+)$", href)
        if match:
            return int(match.group(1))
    return None


def _resolve_metadata() -> tuple[str, str, list[dict[str, object]]]:
    dataset_key = quote(f"doi:{DOI}", safe="")
    dataset_url = f"{API_ROOT}/datasets/{dataset_key}"
    dataset = _json(dataset_url)

    version_url = _href(dataset, "stash:version") or _href(dataset, "version")
    if version_url is None:
        versions = _json(f"{dataset_url}/versions")
        embedded = versions.get("_embedded", {}).get("stash:versions", [])
        if not embedded:
            raise RuntimeError("Dryad metadata exposed no public version")
        first = embedded[0]
        version_url = _href(first, "self")
        if version_url is None and first.get("id") is not None:
            version_url = f"{API_ROOT}/versions/{first['id']}"
    if version_url is None:
        raise RuntimeError("could not resolve Dryad version")

    version = _json(version_url)
    files_url = _href(version, "stash:files") or _href(version, "files")
    if files_url is None:
        version_id = version.get("id")
        if version_id is None:
            match = re.search(r"/versions/(\d+)$", version_url)
            version_id = int(match.group(1)) if match else None
        if version_id is None:
            raise RuntimeError("could not resolve Dryad version id")
        files_url = f"{API_ROOT}/versions/{version_id}/files"

    page = _json(files_url)
    items = page.get("_embedded", {}).get("stash:files", page.get("files", []))
    if not items:
        raise RuntimeError("Dryad version returned no files")
    by_name = {str(item.get("path") or item.get("name") or ""): item for item in items}

    resolved: list[dict[str, object]] = []
    for filename, role in FILES:
        item = by_name.get(filename)
        if item is None:
            raise RuntimeError(f"locked source file absent from Dryad metadata: {filename}")
        file_id = _file_id(item)
        if file_id is None:
            raise RuntimeError(f"could not resolve file id for {filename}")
        resolved.append(
            {
                "filename": filename,
                "role": role,
                "file_id": file_id,
                "metadata_size": item.get("size"),
                "metadata_digest": item.get("digest"),
                "metadata_digest_type": item.get("digestType"),
                "metadata_download_relation": _href(item, "stash:download") or _href(item, "download"),
            }
        )
    return dataset_url, files_url, resolved


def _download_file(meta: dict[str, object]) -> tuple[str, bytes, list[dict[str, object]]]:
    file_id = int(meta["file_id"])
    candidates: list[str] = []
    relation = meta.get("metadata_download_relation")
    if relation:
        candidates.append(str(relation))
    candidates.append(f"{DRYAD_ROOT}/stash/downloads/file_stream/{file_id}")

    attempts: list[dict[str, object]] = []
    seen: set[str] = set()
    for url in candidates:
        if url in seen:
            continue
        seen.add(url)
        try:
            payload = _request(url, accept="application/octet-stream,*/*")
            attempts.append({"url": url, "status": "success", "bytes": len(payload)})
        except HTTPError as exc:
            attempts.append({"url": url, "status": f"http_{exc.code}"})
            continue
        except Exception as exc:
            attempts.append({"url": url, "status": f"error:{type(exc).__name__}"})
            continue

        metadata_size = meta.get("metadata_size")
        if metadata_size is not None and int(metadata_size) != len(payload):
            attempts[-1]["status"] = "size_mismatch"
            attempts[-1]["metadata_size"] = int(metadata_size)
            continue
        return url, payload, attempts

    raise RuntimeError(f"no validated public file payload for {meta['filename']}: {attempts!r}")


def _xlsx_schema(payload: bytes) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    out: list[dict[str, object]] = []
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        out.append(
            {
                "sheet": ws.title,
                "rows_including_header": int(ws.max_row or 0),
                "columns_count": int(ws.max_column or 0),
                "columns": ["" if value is None else str(value) for value in first],
            }
        )
    wb.close()
    return out


def _xls_schema(payload: bytes) -> list[dict[str, object]]:
    wb = xlrd.open_workbook(file_contents=payload, on_demand=True)
    out: list[dict[str, object]] = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        columns = ["" if value is None else str(value) for value in (ws.row_values(0) if ws.nrows else [])]
        out.append(
            {
                "sheet": name,
                "rows_including_header": int(ws.nrows),
                "columns_count": int(ws.ncols),
                "columns": columns,
            }
        )
    wb.release_resources()
    return out


def discover(manifest_path: Path) -> dict[str, object]:
    dataset_url, files_url, files = _resolve_metadata()
    inventory: list[dict[str, object]] = []
    for meta in files:
        url, payload, attempts = _download_file(meta)
        filename = str(meta["filename"])
        schema = _xls_schema(payload) if filename.lower().endswith(".xls") else _xlsx_schema(payload)
        inventory.append(
            {
                **meta,
                "resolved_download_url": url,
                "access_attempts": attempts,
                "observed_bytes": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
                "workbook_schema": schema,
            }
        )

    result: dict[str, object] = {
        "status": "schema_only_discovery_complete",
        "dataset_doi": DOI,
        "dataset_metadata_url": dataset_url,
        "files_metadata_url": files_url,
        "files": inventory,
        "inspection_boundary": (
            "Only Dryad metadata, exact filenames/file ids, file hashes/sizes, workbook sheet names/dimensions and first-row column labels were inspected. "
            "No data row, visitation value, fruit-set value, paternity/genotype value, effect direction, coefficient, p-value or descriptive outcome statistic was read or computed."
        ),
        "next_gate": (
            "Map garden/time/plant/compatibility keys and I_realised/F_reproduction/G_mating-C/R roles from header labels only; "
            "classify direct_joint_state_identifiable, direct_partial_state_identifiable, or not_identifiable_from_archive. "
            "If identifiable, commit a second exact-model preregistration before any outcome row is inspected."
        ),
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", default="artifacts/empirical/witheringia_direct_interaction_schema.json")
    args = parser.parse_args()
    result = discover(Path(args.manifest))
    print(
        json.dumps(
            {
                "status": result["status"],
                "files": [
                    {
                        "filename": row["filename"],
                        "file_id": row["file_id"],
                        "sheets": [sheet["sheet"] for sheet in row["workbook_schema"]],
                    }
                    for row in result["files"]
                ],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
