from __future__ import annotations

import argparse
import hashlib
import io
import json
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

import numpy as np
import openpyxl

RECORD_ID = 4969330
DRYAD_DOI = "10.5061/dryad.5nj81nf"
FILE_NAME = "Koski et al. 2018_Data_ProcRoySoc.xlsx"
EXPECTED_MD5 = "2d26307743e8a22384781854b8f2f33b"
EXPECTED_SHA256 = "b81b77248b75330049e1ddd8ae026db127f838e979620a0415a5addb9a7e8f27"
SHEET = "PopVis Rates_ PL_Depletion"
USER_AGENT = "eco-genetic-warning-extensions/1.0"
EXPECTED_ROWS = 23
TOL = 1e-10

PAIRS = (
    ("Bumble Female Rate", "Bumble Grains Dep Per Hour"),
    ("Megachile Female Rate", "Mega Grains Dep Per Hour"),
    ("Small Female Rate", "Small Grains Dep Per Hour"),
    ("Bumble Male Rate", "Bumble Grains Rem Per Hour"),
    ("Megachile Male Rate", "Mega Grains Rem Per Hour"),
    ("Small Male Rate", "Small Grains Rem Per Hour"),
)
PREDICTOR_COLUMNS = tuple(dict.fromkeys(column for pair in PAIRS for column in pair))


def _download() -> tuple[str, bytes]:
    encoded = quote(FILE_NAME)
    url = f"https://zenodo.org/records/{RECORD_ID}/files/{encoded}?download=1"
    request = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream,*/*"
            ),
        },
    )
    with urlopen(request, timeout=180) as response:
        payload = response.read()
    observed_md5 = hashlib.md5(payload).hexdigest()
    observed_sha256 = hashlib.sha256(payload).hexdigest()
    if observed_md5 != EXPECTED_MD5:
        raise RuntimeError(f"source MD5 mismatch: expected={EXPECTED_MD5}, observed={observed_md5}")
    if observed_sha256 != EXPECTED_SHA256:
        raise RuntimeError(
            f"source SHA256 mismatch: expected={EXPECTED_SHA256}, observed={observed_sha256}"
        )
    return url, payload


def _load_predictors(payload: bytes) -> dict[str, np.ndarray]:
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"locked sheet absent: {SHEET}")
    sheet = workbook[SHEET]

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    header_map = {
        str(cell.value).strip(): cell.column
        for cell in header_cells
        if cell.value is not None and str(cell.value).strip()
    }
    missing = [column for column in PREDICTOR_COLUMNS if column not in header_map]
    if missing:
        workbook.close()
        raise RuntimeError(f"locked predictor columns absent: {missing}")

    values: dict[str, list[float]] = {column: [] for column in PREDICTOR_COLUMNS}
    for row_index in range(2, 2 + EXPECTED_ROWS):
        for column in PREDICTOR_COLUMNS:
            value = sheet.cell(row=row_index, column=header_map[column]).value
            try:
                numeric = float(value)
            except (TypeError, ValueError) as exc:
                workbook.close()
                raise RuntimeError(
                    f"non-numeric predictor: row={row_index}, column={column!r}, value={value!r}"
                ) from exc
            if not np.isfinite(numeric):
                workbook.close()
                raise RuntimeError(
                    f"non-finite predictor: row={row_index}, column={column!r}, value={numeric!r}"
                )
            values[column].append(numeric)

    # The source sheet was locked by #114 to exactly 23 population rows. Check
    # that no additional nonblank predictor-bearing population row exists.
    next_row = 2 + EXPECTED_ROWS
    if any(sheet.cell(row=next_row, column=header_map[column]).value not in (None, "") for column in PREDICTOR_COLUMNS):
        workbook.close()
        raise RuntimeError("more than 23 predictor-bearing rows detected")
    workbook.close()

    arrays = {column: np.asarray(series, dtype=float) for column, series in values.items()}
    if any(array.size != EXPECTED_ROWS for array in arrays.values()):
        raise RuntimeError("predictor row count mismatch")
    return arrays


def _zscore(values: np.ndarray) -> np.ndarray:
    sd = float(np.std(values, ddof=0))
    if sd <= 0 or not np.isfinite(sd):
        raise RuntimeError("predictor has zero or non-finite standard deviation")
    return (values - float(np.mean(values))) / sd


def _diagnose_pair(x: np.ndarray, y: np.ndarray) -> dict[str, object]:
    nonzero = x != 0.0
    ratios = y[nonzero] / x[nonzero]
    if ratios.size < 2 or not np.isfinite(ratios).all():
        raise RuntimeError("fewer than two finite nonzero-source ratios")

    median_ratio = float(np.median(ratios))
    denominator = max(abs(median_ratio), 1e-15)
    max_relative_deviation = float(np.max(np.abs(ratios - median_ratio) / denominator))
    zero_source_effective_values = y[~nonzero]
    zero_rows_match = bool(np.all(zero_source_effective_values == 0.0))
    max_zscore_difference = float(np.max(np.abs(_zscore(x) - _zscore(y))))

    confirmed = bool(
        median_ratio > 0
        and max_relative_deviation <= TOL
        and zero_rows_match
        and max_zscore_difference <= TOL
    )
    return {
        "n_rows": int(x.size),
        "n_nonzero_source_rows": int(ratios.size),
        "n_zero_source_rows": int((~nonzero).sum()),
        "median_ratio": median_ratio,
        "min_ratio": float(np.min(ratios)),
        "max_ratio": float(np.max(ratios)),
        "max_relative_ratio_deviation": max_relative_deviation,
        "zero_source_rows_have_zero_effective": zero_rows_match,
        "max_abs_zscore_difference": max_zscore_difference,
        "tolerance": TOL,
        "decision": "constant_rescaling_confirmed" if confirmed else "not_constant_rescaling",
    }


def run() -> dict[str, object]:
    url, payload = _download()
    predictors = _load_predictors(payload)
    pair_results: list[dict[str, object]] = []
    for phase_column, effective_column in PAIRS:
        diagnostic = _diagnose_pair(predictors[phase_column], predictors[effective_column])
        pair_results.append(
            {
                "phase_column": phase_column,
                "effective_column": effective_column,
                **diagnostic,
            }
        )

    all_confirmed = all(
        result["decision"] == "constant_rescaling_confirmed" for result in pair_results
    )
    decision = "constant_rescaling_confirmed" if all_confirmed else "not_constant_rescaling"
    return {
        "stage": "Campanula predictor-only effective-interaction rescaling diagnostic",
        "decision": decision,
        "source_lock": {
            "dryad_doi": DRYAD_DOI,
            "zenodo_record": RECORD_ID,
            "filename": FILE_NAME,
            "download_url": url,
            "md5": hashlib.md5(payload).hexdigest(),
            "sha256": hashlib.sha256(payload).hexdigest(),
            "sheet": SHEET,
        },
        "sample": {"n_rows": EXPECTED_ROWS},
        "pairs": pair_results,
        "response_firewall": (
            "Numeric reads were restricted to the twelve preregistered predictor columns; "
            "no outcome column values were read or used."
        ),
        "interpretation_boundary": (
            "This diagnostic can explain algebraic collapse under feature-wise standardization. "
            "It cannot change the locked #114 predictive decision or establish biological irrelevance of efficiency."
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        default="artifacts/empirical/campanula_rescaling_diagnostic.json",
    )
    args = parser.parse_args()
    try:
        result = run()
    except Exception as exc:
        result = {
            "stage": "Campanula predictor-only effective-interaction rescaling diagnostic",
            "decision": "not_identifiable",
            "reason": f"{type(exc).__name__}: {exc}",
            "response_firewall": "No outcome model was run in this diagnostic.",
        }
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({
        "decision": result.get("decision"),
        "reason": result.get("reason"),
        "pairs": result.get("pairs"),
    }, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
