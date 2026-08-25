from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

import openpyxl

RECORD_ID = 4969330
DRYAD_DOI = "10.5061/dryad.5nj81nf"
FILE_NAME = "Koski et al. 2018_Data_ProcRoySoc.xlsx"
EXPECTED_MD5 = "2d26307743e8a22384781854b8f2f33b"
USER_AGENT = "eco-genetic-warning-extensions/1.0"
TOP_ROWS = 10
METADATA_ROWS = 100

KEYWORDS = {
    "population": ("population", "pop"),
    "pollen_limitation": ("pollen limitation", "pollen.limit", "pl"),
    "visitation": ("visitation", "visit rate", "visits", "visitor"),
    "pollinator_group": ("pollinator", "bee", "bumble", "megachile", "solitary"),
    "pollen_deposition": ("pollen deposition", "pollen deposited", "stig", "deposition"),
    "seed_set": ("seed set", "seedset", "seed"),
    "pollen_removal": ("pollen removal", "pollen removed", "removal"),
    "efficiency": ("efficiency", "single visit", "single-visit", "per visit"),
}


def _download() -> tuple[str, bytes]:
    encoded = quote(FILE_NAME)
    url = f"https://zenodo.org/records/{RECORD_ID}/files/{encoded}?download=1"
    req = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        },
    )
    with urlopen(req, timeout=180) as response:
        payload = response.read()
    observed = hashlib.md5(payload).hexdigest()
    if observed != EXPECTED_MD5:
        raise RuntimeError(f"source MD5 mismatch: expected={EXPECTED_MD5}, observed={observed}")
    return url, payload


def _safe_string(value: object) -> str | None:
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return None


def _row_type_signature(row) -> list[str]:
    signature: list[str] = []
    for cell in row:
        value = cell.value
        if value is None:
            signature.append("blank")
        elif isinstance(value, str):
            signature.append("string")
        elif isinstance(value, bool):
            signature.append("boolean")
        elif isinstance(value, (int, float)):
            signature.append("numeric")
        else:
            signature.append(type(value).__name__)
    return signature


def _sheet_schema(ws) -> dict[str, object]:
    top_rows: list[dict[str, object]] = []
    strings_for_keywords: list[str] = []

    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(TOP_ROWS, ws.max_row)), start=1):
        string_cells: list[dict[str, object]] = []
        for column_index, cell in enumerate(row, start=1):
            text = _safe_string(cell.value)
            if text is not None:
                strings_for_keywords.append(text)
                string_cells.append({"column": column_index, "text": text})
        top_rows.append(
            {
                "row": row_index,
                "type_signature": _row_type_signature(row),
                "string_cells": string_cells,
            }
        )

    metadata_text: list[dict[str, object]] = []
    if "metadata" in ws.title.lower():
        for row_index, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(METADATA_ROWS, ws.max_row)), start=1
        ):
            for column_index, cell in enumerate(row, start=1):
                text = _safe_string(cell.value)
                if text is not None:
                    strings_for_keywords.append(text)
                    metadata_text.append(
                        {"row": row_index, "column": column_index, "text": text}
                    )

    lowered = "\n".join(strings_for_keywords).lower()
    keyword_hits = {
        label: sorted({term for term in terms if term in lowered})
        for label, terms in KEYWORDS.items()
    }

    return {
        "sheet": ws.title,
        "rows": int(ws.max_row),
        "columns": int(ws.max_column),
        "top_rows": top_rows,
        "metadata_string_cells": metadata_text,
        "label_keyword_hits": keyword_hits,
    }


def discover(output: Path) -> dict[str, object]:
    url, payload = _download()
    temporary = output.parent / "_campanula_effective_interaction_locked.xlsx"
    temporary.parent.mkdir(parents=True, exist_ok=True)
    temporary.write_bytes(payload)

    workbook = openpyxl.load_workbook(temporary, read_only=True, data_only=False)
    try:
        sheets = [_sheet_schema(ws) for ws in workbook.worksheets]
    finally:
        workbook.close()
        temporary.unlink(missing_ok=True)

    result: dict[str, object] = {
        "status": "schema_only_discovery_complete",
        "source_lock": {
            "zenodo_record": RECORD_ID,
            "dryad_doi": DRYAD_DOI,
            "filename": FILE_NAME,
            "download_url": url,
            "published_md5": EXPECTED_MD5,
            "observed_md5": hashlib.md5(payload).hexdigest(),
            "sha256": hashlib.sha256(payload).hexdigest(),
            "bytes": len(payload),
        },
        "workbook": {
            "sheet_names": [sheet["sheet"] for sheet in sheets],
            "sheet_count": len(sheets),
            "sheets": sheets,
        },
        "inspection_boundary": (
            "Only workbook structure, cell types in the first ten rows, string labels in the first ten rows, "
            "and string-only text from an explicitly named metadata sheet were retained. Numeric study-cell values "
            "were not copied, summarized, compared, modelled, or used to choose an interaction state."
        ),
        "next_gate": (
            "Using labels/text only, classify effective_interaction_state_identifiable, "
            "partial_effective_interaction_state_identifiable, or not_identifiable_from_archive. "
            "If identifiable, commit a second exact-model preregistration before any numeric study value is inspected."
        ),
    }
    output.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        default="artifacts/empirical/campanula_effective_interaction_schema.json",
    )
    args = parser.parse_args()
    result = discover(Path(args.output))
    print(
        json.dumps(
            {
                "status": result["status"],
                "sheets": result["workbook"]["sheet_names"],
                "sha256": result["source_lock"]["sha256"],
            },
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
