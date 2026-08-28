from __future__ import annotations

import argparse
import hashlib
import io
import json
import math
import re
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from scipy.special import gammaln

DOI = "10.5061/dryad.hqbzkh1bm"
NETWORK_SHEETS = {"2016": "Sheet1_Network2016", "2017": "Sheet2_Network2017"}
SEM_SHEET = "Sheet 3_SEMvariables"
EXPECTED_XLSX_BYTES = 60483
EXPECTED_XLSX_SHA256 = "a6490b3699468898e99bb65b528d67e01d876dc85aec2847edf8191b806cb56f"
BOOTSTRAP_SEED = 20260828
BOOTSTRAP_N = 10000


def _req(url: str, json_only: bool = False) -> urllib.request.Request:
    return urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
            "Accept": "application/json" if json_only else "application/zip,application/octet-stream,*/*;q=0.8",
        },
    )


def _json(url: str) -> dict[str, Any]:
    with urllib.request.urlopen(_req(url, json_only=True), timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def _norm_species(value: object) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _year(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def download_locked_workbook() -> tuple[bytes, dict[str, Any]]:
    encoded = urllib.parse.quote(f"doi:{DOI}", safe="")
    versions = _json(f"https://datadryad.org/api/v2/datasets/{encoded}/versions").get("_embedded", {}).get("stash:versions", [])
    if not versions:
        raise RuntimeError("no Dryad versions")
    latest = versions[-1]
    download_href = latest["_links"]["stash:download"]["href"]
    download_url = "https://datadryad.org" + download_href if str(download_href).startswith("/") else str(download_href)
    with urllib.request.urlopen(_req(download_url), timeout=120) as response:
        bundle = response.read()
        final_url = response.geturl()
    with zipfile.ZipFile(io.BytesIO(bundle)) as zf:
        by_name = {Path(name).name: name for name in zf.namelist()}
        data = zf.read(by_name["Dryad.xlsx"])
    digest = hashlib.sha256(data).hexdigest()
    if len(data) != EXPECTED_XLSX_BYTES or digest != EXPECTED_XLSX_SHA256:
        raise RuntimeError("locked Dryad.xlsx identity mismatch")
    return data, {
        "version": latest["_links"]["self"]["href"],
        "download_final_url": final_url,
        "bytes": len(data),
        "sha256": digest,
    }


def _numeric_or_zero(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float, np.integer, np.floating)):
        x = float(value)
        if not math.isfinite(x):
            raise RuntimeError("nonfinite network link weight")
        return x
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        x = float(text)
    except ValueError as exc:
        raise RuntimeError(f"nonnumeric nonblank network link weight: {text!r}") from exc
    if not math.isfinite(x):
        raise RuntimeError("nonfinite network link weight")
    return x


def build_frame(workbook_bytes: bytes) -> tuple[pd.DataFrame, dict[str, Any]]:
    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)

    visit_rows: list[dict[str, Any]] = []
    for year, sheet_name in NETWORK_SHEETS.items():
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        pollinator_count = sum(1 for value in header[1:] if str(value or "").strip())
        for row in rows:
            species = _norm_species(row[0] if row else None)
            if not species:
                continue
            weights = [_numeric_or_zero(value) for value in row[1:]]
            visit_rows.append(
                {
                    "Year": year,
                    "SpeciesNorm": species,
                    "I_visit": float(sum(weights)),
                    "network_pollinator_columns": pollinator_count,
                }
            )
    visit = pd.DataFrame(visit_rows)
    if visit.duplicated(["Year", "SpeciesNorm"]).any():
        raise RuntimeError("duplicate plant row in year-specific network")

    ws = wb[SEM_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = [str(value or "").strip() for value in next(rows)]
    required = ["Year", "Species", "DPD", "FloralUnitSize", "SeedsFlowerRounded"]
    missing = [name for name in required if name not in header]
    if missing:
        raise RuntimeError(f"missing SEM columns: {missing}")
    idx = {name: header.index(name) for name in required}
    sem_rows = []
    for row in rows:
        year = _year(row[idx["Year"]])
        species = _norm_species(row[idx["Species"]])
        if year not in NETWORK_SHEETS or not species:
            continue
        sem_rows.append(
            {
                "Year": year,
                "SpeciesNorm": species,
                "DPD": row[idx["DPD"]],
                "FloralUnitSize": row[idx["FloralUnitSize"]],
                "SeedsFlowerRounded": row[idx["SeedsFlowerRounded"]],
            }
        )
    sem = pd.DataFrame(sem_rows)
    if sem.duplicated(["Year", "SpeciesNorm"]).any():
        raise RuntimeError("duplicate SEM species-year row")

    frame = sem.merge(visit, on=["Year", "SpeciesNorm"], how="inner", validate="one_to_one")
    before = len(frame)
    for column in ["DPD", "FloralUnitSize", "SeedsFlowerRounded", "I_visit"]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame = frame.dropna(subset=["DPD", "FloralUnitSize", "SeedsFlowerRounded", "I_visit"]).copy()
    frame = frame[np.isfinite(frame[["DPD", "FloralUnitSize", "SeedsFlowerRounded", "I_visit"]]).all(axis=1)].copy()
    frame = frame[frame["SeedsFlowerRounded"] >= 0].copy()
    if len(frame) != 41:
        raise RuntimeError(f"eligible row count drifted from locked Stage A: {len(frame)} != 41")
    species_count = int(frame["SpeciesNorm"].nunique())
    if species_count != 22:
        raise RuntimeError(f"eligible species count drifted from locked Stage A: {species_count} != 22")
    year_counts = {str(k): int(v) for k, v in frame["Year"].value_counts().sort_index().items()}
    if year_counts != {"2016": 21, "2017": 20}:
        raise RuntimeError(f"eligible year counts drifted from locked Stage A: {year_counts}")
    audit = {
        "merged_rows_before_numeric_checks": before,
        "eligible_rows": len(frame),
        "distinct_species": species_count,
        "year_counts": year_counts,
        "zero_visit_rows": int((frame["I_visit"] == 0).sum()),
    }
    return frame.sort_values(["SpeciesNorm", "Year"]).reset_index(drop=True), audit


def _design(train: pd.DataFrame, test: pd.DataFrame, include_process: bool) -> tuple[np.ndarray, np.ndarray, list[str]]:
    train_parts = [np.ones((len(train), 1), dtype=float)]
    test_parts = [np.ones((len(test), 1), dtype=float)]
    columns = ["intercept"]

    columns.append("Year[2017]")
    train_parts.append((train["Year"].astype(str).to_numpy() == "2017").astype(float)[:, None])
    test_parts.append((test["Year"].astype(str).to_numpy() == "2017").astype(float)[:, None])

    continuous = ["DPD", "FloralUnitSize"] + (["I_visit"] if include_process else [])
    for name in continuous:
        mean = float(train[name].mean())
        sd = float(train[name].std(ddof=0))
        if not math.isfinite(sd) or sd <= 0:
            raise RuntimeError(f"training-fold predictor has zero/nonfinite SD: {name}")
        columns.append(f"z({name})")
        train_parts.append(((train[name].to_numpy(dtype=float) - mean) / sd)[:, None])
        test_parts.append(((test[name].to_numpy(dtype=float) - mean) / sd)[:, None])
    return np.hstack(train_parts), np.hstack(test_parts), columns


def _poisson_nll(y: np.ndarray, mu: np.ndarray) -> np.ndarray:
    if np.any(mu <= 0) or not np.all(np.isfinite(mu)):
        raise RuntimeError("nonpositive/nonfinite Poisson prediction")
    return mu - y * np.log(mu) + gammaln(y + 1.0)


def _fit_score(train: pd.DataFrame, test: pd.DataFrame, include_process: bool) -> tuple[np.ndarray, list[str]]:
    x_train, x_test, columns = _design(train, test, include_process)
    y_train = train["SeedsFlowerRounded"].to_numpy(dtype=float)
    y_test = test["SeedsFlowerRounded"].to_numpy(dtype=float)
    model = sm.GLM(y_train, x_train, family=sm.families.Poisson())
    fit = model.fit(maxiter=200, disp=0)
    if not bool(fit.converged):
        raise RuntimeError("Poisson GLM did not converge")
    mu = np.asarray(fit.predict(x_test), dtype=float)
    return _poisson_nll(y_test, mu), columns


def run_loso(frame: pd.DataFrame) -> dict[str, Any]:
    records = []
    try:
        for species in sorted(frame["SpeciesNorm"].unique()):
            test = frame[frame["SpeciesNorm"] == species].copy()
            train = frame[frame["SpeciesNorm"] != species].copy()
            nll0, cols0 = _fit_score(train, test, include_process=False)
            nll1, cols1 = _fit_score(train, test, include_process=True)
            records.append(
                {
                    "species": species,
                    "species_year_rows": int(len(test)),
                    "NLL_M0": float(nll0.sum()),
                    "NLL_M1": float(nll1.sum()),
                    "delta_NLL": float(nll1.sum() - nll0.sum()),
                }
            )
    except Exception as exc:
        return {
            "decision": "primary_model_not_identifiable",
            "error": f"{type(exc).__name__}: {exc}",
            "species_folds_completed": len(records),
            "species_records": records,
        }

    deltas = np.asarray([record["delta_NLL"] for record in records], dtype=float)
    total = float(deltas.sum())
    rng = np.random.default_rng(BOOTSTRAP_SEED)
    draws = rng.choice(deltas, size=(BOOTSTRAP_N, len(deltas)), replace=True).sum(axis=1)
    lo, hi = np.quantile(draws, [0.025, 0.975])
    decision = "process_information_detected" if total < 0 and float(hi) < 0 else "no_detected_process_information"
    return {
        "decision": decision,
        "delta_NLL_total_M1_minus_M0": total,
        "species_bootstrap_95ci": [float(lo), float(hi)],
        "bootstrap_seed": BOOTSTRAP_SEED,
        "bootstrap_n": BOOTSTRAP_N,
        "M0_columns": cols0,
        "M1_columns": cols1,
        "species_records": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="artifacts/empirical/n3_mallorca_network_process_adequacy_result.json")
    args = parser.parse_args()
    result: dict[str, Any] = {
        "analysis": "N3_Mallorca_network_process_adequacy",
        "doi": DOI,
        "response_firewall": "Stage A and the B1 endpoint, process definition, baseline terms, Poisson family, LOSO holdout, score, species bootstrap and no-rescue rules were committed before project-computed visitation row sums were associated with SeedsFlowerRounded values.",
    }
    try:
        workbook, provenance = download_locked_workbook()
        result["source_provenance"] = provenance
        frame, audit = build_frame(workbook)
        result["data_audit"] = audit
        result.update(run_loso(frame))
    except Exception as exc:
        result["decision"] = "primary_model_not_identifiable"
        result["error"] = f"{type(exc).__name__}: {exc}"
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({
        "decision": result["decision"],
        "delta_NLL_total": result.get("delta_NLL_total_M1_minus_M0"),
        "species_bootstrap_95ci": result.get("species_bootstrap_95ci"),
    }, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
