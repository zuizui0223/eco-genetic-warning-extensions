from __future__ import annotations

import argparse
import csv
import hashlib
import html as html_module
import io
import json
import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

import openpyxl
import xlrd

RECORDS = (
    {
        "role": "wild_visitation",
        "record_id": 5552,
        "doi": "10.15479/AT:ISTA:36",
        "filename": "IST-2016-36-v1+1_tag_assay_archive.zip",
        "md5": "cbc61b523d4d475a04a737d50dc470ef",
    },
    {
        "role": "wild_paternity_2012",
        "record_id": 5553,
        "doi": "10.15479/AT:ISTA:37",
        "filename": "IST-2016-37-v1+1_paternity_archive.zip",
        "md5": "4ae751b1fa4897fa216241f975a57313",
    },
)
ROOT = "https://research-explorer.ista.ac.at"
OAI_ROOT = "https://research-explorer.app.ist.ac.at/oai"
UA = "eco-genetic-warning-extensions/1.0"


class MetadataParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.values: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        for _, value in attrs:
            if value:
                self.values.append(value)


def _request(url: str, *, accept: str) -> bytes:
    req = Request(url, headers={"User-Agent": UA, "Accept": accept})
    with urlopen(req, timeout=180) as response:
        return response.read()


def _extract_exact_url_candidates(text: str, base_url: str, exact: str) -> list[str]:
    decoded = html_module.unescape(text).replace("\\/", "/")
    parser = MetadataParser()
    try:
        parser.feed(decoded)
    except Exception:
        pass

    values = list(parser.values)
    values.extend(re.findall(r"https?://[^\s\"'<>]+", decoded))
    values.extend(re.findall(r"(?:href|url|download|file)\s*[:=]\s*[\"']([^\"']+)[\"']", decoded, re.I))

    candidates: set[str] = set()
    for value in values:
        value = value.strip()
        if exact not in value:
            continue
        url = urljoin(base_url, value)
        parsed = urlparse(url)
        if parsed.scheme == "https" and (parsed.netloc.endswith("ista.ac.at") or parsed.netloc.endswith("ist.ac.at")):
            candidates.add(url)
    return sorted(candidates)


def _oai_url(record_id: int) -> str:
    identifier = f"oai:pub.research-explorer.ista.ac.at:{record_id}"
    return f"{OAI_ROOT}?verb=GetRecord&metadataPrefix=oai_dc&identifier={identifier}"


def _resolve_archive(record: dict[str, object]) -> tuple[str, bytes, list[dict[str, object]]]:
    landing = f"{ROOT}/record/{record['record_id']}"
    exact = str(record["filename"])
    attempts: list[dict[str, object]] = []
    candidates: set[str] = set()

    # Attempt 1 showed that the visible file name is not represented as a
    # normal <a href>. The second resolver therefore uses only embedded landing
    # metadata plus the OAI metadata endpoint independently advertised for the
    # same ISTA record by B2FIND. It never guesses a download route.
    landing_html = _request(landing, accept="text/html,*/*")
    landing_text = landing_html.decode("utf-8", errors="replace")
    candidates.update(_extract_exact_url_candidates(landing_text, landing, exact))
    attempts.append({"metadata_source": landing, "exact_url_candidates": len(candidates)})

    oai = _oai_url(int(record["record_id"]))
    try:
        oai_payload = _request(oai, accept="application/xml,text/xml,*/*")
        oai_text = oai_payload.decode("utf-8", errors="replace")
        before = len(candidates)
        candidates.update(_extract_exact_url_candidates(oai_text, oai, exact))
        attempts.append({"metadata_source": oai, "exact_url_candidates_added": len(candidates) - before})
    except Exception as exc:
        attempts.append({"metadata_source": oai, "status": f"error:{type(exc).__name__}"})

    if not candidates:
        raise RuntimeError(f"ISTA public landing/OAI metadata exposed no exact archive URL for {exact}; attempts={attempts!r}")

    failures: list[str] = []
    for url in sorted(candidates):
        try:
            payload = _request(url, accept="application/zip,application/octet-stream,*/*")
        except Exception as exc:
            failures.append(f"{url}:{type(exc).__name__}")
            continue
        observed = hashlib.md5(payload).hexdigest()
        if observed != record["md5"]:
            failures.append(f"{url}:md5={observed}")
            continue
        if not zipfile.is_zipfile(io.BytesIO(payload)):
            failures.append(f"{url}:not_zip")
            continue
        return url, payload, attempts
    raise RuntimeError(f"no exact MD5-locked archive resolved for {exact}: candidates={sorted(candidates)!r}, failures={failures!r}")


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise RuntimeError("could not decode text header")


def _delimited_header(raw: bytes, suffix: str) -> list[str]:
    text = _decode_text(raw)
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = "\t" if suffix in {".tsv", ".tab"} else ","
    if suffix == ".txt":
        delimiter = "\t" if "\t" in first_line else ","
    row = next(csv.reader([first_line], delimiter=delimiter), [])
    return [str(v) for v in row]


def _xlsx_schema(raw: bytes) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
    out: list[dict[str, object]] = []
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        out.append({
            "sheet": ws.title,
            "rows_including_header": int(ws.max_row or 0),
            "columns_count": int(ws.max_column or 0),
            "columns": ["" if x is None else str(x) for x in first],
        })
    wb.close()
    return out


def _xls_schema(raw: bytes) -> list[dict[str, object]]:
    wb = xlrd.open_workbook(file_contents=raw, on_demand=True)
    out: list[dict[str, object]] = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        out.append({
            "sheet": name,
            "rows_including_header": int(ws.nrows),
            "columns_count": int(ws.ncols),
            "columns": [str(v) for v in (ws.row_values(0) if ws.nrows else [])],
        })
    wb.release_resources()
    return out


def _member_schema(name: str, raw: bytes) -> dict[str, object]:
    suffix = Path(name).suffix.lower()
    result: dict[str, object] = {
        "member_name": name,
        "member_bytes": len(raw),
        "sha256": hashlib.sha256(raw).hexdigest(),
    }
    if suffix in {".csv", ".tsv", ".tab", ".txt"}:
        result["columns"] = _delimited_header(raw, suffix)
    elif suffix == ".xlsx":
        result["workbook_schema"] = _xlsx_schema(raw)
    elif suffix == ".xls":
        result["workbook_schema"] = _xls_schema(raw)
    return result


def discover(manifest_path: Path) -> dict[str, object]:
    records: list[dict[str, object]] = []
    for record in RECORDS:
        url, payload, resolution_attempts = _resolve_archive(record)
        members: list[dict[str, object]] = []
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                raw = archive.read(info)
                members.append(_member_schema(info.filename, raw))
        records.append({
            **record,
            "resolved_archive_url": url,
            "resolution_attempts": resolution_attempts,
            "archive_bytes": len(payload),
            "observed_md5": hashlib.md5(payload).hexdigest(),
            "sha256": hashlib.sha256(payload).hexdigest(),
            "members": members,
        })

    result: dict[str, object] = {
        "status": "schema_only_discovery_complete",
        "records": records,
        "inspection_boundary": (
            "Only ISTA landing/OAI metadata, exact MD5-locked archive bytes, archive member names/sizes/hashes, and first header rows or workbook header rows were inspected. "
            "No data-row value, visitation outcome, reproductive value, paternity assignment, genotype value, effect direction, coefficient, p-value or descriptive outcome statistic was read or computed."
        ),
        "next_gate": (
            "Use schema/header names only to map 2012 year/time, wild plant identity, spatial/phenotypic state, I_realised_proxy, any explicit F_reproduction endpoint, and G_mating/C_pollen. "
            "Classify wild_IFG_joint_state_identifiable, wild_IG_partial_state_identifiable, or wild_state_not_identifiable."
        ),
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", default="artifacts/empirical/antirrhinum_realised_visitation_schema.json")
    args = parser.parse_args()
    result = discover(Path(args.manifest))
    print(json.dumps({
        "status": result["status"],
        "records": {
            row["role"]: [member["member_name"] for member in row["members"]]
            for row in result["records"]
        },
    }, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
