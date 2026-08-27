from __future__ import annotations

import argparse
import hashlib
import io
import json
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

DOI = "10.5061/dryad.hqbzkh1bm"
REQUIRED_FILES = {"Dryad.xlsx", "Dryad_notes.docx"}


def _request(url: str, *, json_only: bool = False) -> urllib.request.Request:
    return urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
            "Accept": "application/json" if json_only else "application/zip,application/octet-stream,*/*;q=0.8",
        },
    )


def _json(url: str) -> dict:
    with urllib.request.urlopen(_request(url, json_only=True), timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def _docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for paragraph in root.findall(".//w:p", ns):
        parts = [node.text or "" for node in paragraph.findall(".//w:t", ns)]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _string_schema_rows(ws, limit: int = 12) -> list[dict]:
    out = []
    for row_i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, limit), values_only=True), start=1):
        strings = []
        for col_i, value in enumerate(row, start=1):
            if isinstance(value, str) and value.strip():
                strings.append({"column": col_i, "text": value.strip()})
        if strings:
            out.append({"row": row_i, "string_cells": strings})
    return out


def audit() -> dict:
    encoded = urllib.parse.quote(f"doi:{DOI}", safe="")
    versions = _json(f"https://datadryad.org/api/v2/datasets/{encoded}/versions").get("_embedded", {}).get("stash:versions", [])
    if not versions:
        raise RuntimeError("no Dryad versions")
    latest = versions[-1]
    files_href = latest["_links"]["stash:files"]["href"]
    download_href = latest["_links"]["stash:download"]["href"]
    files_url = "https://datadryad.org" + files_href if files_href.startswith("/") else files_href
    items = _json(files_url).get("_embedded", {}).get("stash:files", [])
    manifest = {
        str(item["path"]): {
            "size": item.get("size"),
            "digest": item.get("digest"),
            "digest_type": item.get("digestType"),
        }
        for item in items
    }
    missing = sorted(REQUIRED_FILES - set(manifest))
    if missing:
        return {"analysis": "Mallorca_original_Dryad_schema_audit", "doi": DOI, "decision": "required_files_missing", "missing": missing}

    download_url = "https://datadryad.org" + download_href if download_href.startswith("/") else download_href
    with urllib.request.urlopen(_request(download_url), timeout=120) as response:
        bundle = response.read()
        download = {"status": int(response.status), "bytes": len(bundle), "final_url": response.geturl()}

    payloads = {}
    verification = {}
    with zipfile.ZipFile(io.BytesIO(bundle)) as zf:
        by_name = {Path(name).name: name for name in zf.namelist()}
        for name in REQUIRED_FILES:
            data = zf.read(by_name[name])
            digest = hashlib.sha256(data).hexdigest()
            accepted = len(data) == int(manifest[name]["size"]) and digest == str(manifest[name]["digest"])
            if not accepted:
                raise RuntimeError(f"identity mismatch for {name}")
            payloads[name] = data
            verification[name] = {"bytes": len(data), "sha256": digest, "accepted": True}

    workbook = load_workbook(io.BytesIO(payloads["Dryad.xlsx"]), read_only=True, data_only=False)
    sheets = {}
    for name in workbook.sheetnames:
        ws = workbook[name]
        sheets[name] = {
            "max_row": int(ws.max_row or 0),
            "max_column": int(ws.max_column or 0),
            "first_rows_string_cells_only": _string_schema_rows(ws),
        }

    notes = _docx_text(payloads["Dryad_notes.docx"])
    terms = ["visit", "rate", "pollinator", "species", "year", "seed", "fitness", "flower", "network"]
    relevant_note_lines = [line for line in notes.splitlines() if any(term in line.casefold() for term in terms)]

    combined_text = "\n".join(
        cell["text"]
        for sheet in sheets.values()
        for row in sheet["first_rows_string_cells_only"]
        for cell in row["string_cells"]
    ).casefold() + "\n" + notes.casefold()
    direct_visit_terms = any(term in combined_text for term in ["visitation rate", "visits per flower", "number of visits", "pollinator visits"])
    function_terms = any(term in combined_text for term in ["seeds per flower", "seed weight", "fitness"])

    return {
        "analysis": "Mallorca_original_Dryad_schema_audit",
        "doi": DOI,
        "version": latest["_links"]["self"]["href"],
        "download": download,
        "file_verification": verification,
        "workbook": {"sheetnames": workbook.sheetnames, "sheets": sheets},
        "documentation": {"relevant_note_lines": relevant_note_lines},
        "schema_signals": {
            "direct_visitation_language_present": direct_visit_terms,
            "reproductive_function_language_present": function_terms,
        },
        "decision": "schema_documented_for_manual_gate_review",
        "response_firewall": "Only archive identity, workbook sheet names/dimensions, string-only cells from the first 12 rows, and documentation text were inspected. Numeric visitation and reproductive outcome values were not read into the audit output or modeled.",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="artifacts/empirical/n3_mallorca_original_dryad_schema.json")
    args = parser.parse_args()
    try:
        result = audit()
    except Exception as exc:
        result = {
            "analysis": "Mallorca_original_Dryad_schema_audit",
            "doi": DOI,
            "decision": "schema_audit_error",
            "error": f"{type(exc).__name__}: {exc}",
            "response_firewall": "No numeric visitation or reproductive outcome values were modeled.",
        }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({"decision": result["decision"], "schema_signals": result.get("schema_signals")}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
