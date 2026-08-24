from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import zipfile
from pathlib import Path
from urllib.parse import quote, urljoin
from urllib.request import Request, urlopen

import openpyxl

DOI = "10.5061/dryad.1cm80"
DRYAD_ROOT = "https://datadryad.org"
API_ROOT = f"{DRYAD_ROOT}/api/v2"
FILES = (
    (30526, "correlation_dbh_number of inflorescences.xlsx"),
    (30527, "genotypes_Miconia affinis.xlsx"),
    (30528, "pollen_dispersal_analysis_data.xlsx"),
    (30529, "seed_viability_analysis_data.xlsx"),
)


def _request_bytes(url: str, *, accept: str) -> bytes:
    request = Request(
        url,
        headers={
            "User-Agent": "eco-genetic-warning-extensions/1.0",
            "Accept": accept,
            "X-API-Version": "2.1.0",
        },
    )
    with urlopen(request, timeout=180) as response:
        return response.read()


def _json(url: str) -> dict:
    return json.loads(_request_bytes(url, accept="application/json").decode("utf-8"))


def _absolute(href: str) -> str:
    return urljoin(DRYAD_ROOT, href)


def _href(obj: dict, relation: str) -> str | None:
    value = obj.get("_links", {}).get(relation)
    if isinstance(value, dict) and value.get("href"):
        return _absolute(str(value["href"]))
    return None


def _file_id(item: dict) -> int | None:
    # Current Dryad HAL file records may expose the id only in the self link.
    if item.get("id") is not None:
        return int(item["id"])
    self_href = _href(item, "self")
    if self_href:
        match = re.search(r"/files/(\d+)$", self_href)
        if match:
            return int(match.group(1))
    return None


def _resolve_public_metadata() -> tuple[str, str, list[dict[str, object]]]:
    dataset_key = quote(f"doi:{DOI}", safe="")
    dataset_url = f"{API_ROOT}/datasets/{dataset_key}"
    dataset = _json(dataset_url)

    version_url = _href(dataset, "stash:version") or _href(dataset, "version")
    if version_url is None:
        versions_url = f"{dataset_url}/versions"
        versions = _json(versions_url)
        embedded = versions.get("_embedded", {}).get("stash:versions", [])
        if not embedded:
            raise RuntimeError("Dryad dataset metadata exposed no public versions")
        version = embedded[0]
        version_url = _href(version, "self")
        if version_url is None and version.get("id") is not None:
            version_url = f"{API_ROOT}/versions/{version['id']}"
        if version_url is None:
            raise RuntimeError("could not resolve Dryad version URL")
    version = _json(version_url)

    files_url = _href(version, "stash:files") or _href(version, "files")
    if files_url is None:
        version_id = version.get("id")
        if version_id is None:
            match = re.search(r"/versions/(\d+)$", version_url)
            version_id = int(match.group(1)) if match else None
        if version_id is None:
            raise RuntimeError("could not resolve Dryad version id for files listing")
        files_url = f"{API_ROOT}/versions/{version_id}/files"
    page = _json(files_url)
    items = page.get("_embedded", {}).get("stash:files", page.get("files", []))
    if not items:
        raise RuntimeError("Dryad public version returned no files")

    by_name = {str(item.get("path") or item.get("name") or ""): item for item in items}
    resolved: list[dict[str, object]] = []
    for expected_id, filename in FILES:
        item = by_name.get(filename)
        if item is None:
            raise RuntimeError(f"locked Miconia file absent from Dryad metadata: {filename}")
        observed_id = _file_id(item)
        if observed_id != expected_id:
            raise RuntimeError(
                f"locked Miconia file id changed for {filename}: expected={expected_id}, observed={observed_id}"
            )
        resolved.append(
            {
                "file_stream_id": expected_id,
                "filename": filename,
                "metadata_size": item.get("size"),
                "metadata_digest": item.get("digest"),
                "metadata_digest_type": item.get("digestType"),
                "metadata_self": _href(item, "self"),
                "metadata_download_relation": _href(item, "stash:download") or _href(item, "download"),
            }
        )
    return dataset_url, files_url, resolved


def _download_public_package() -> tuple[str, bytes]:
    dataset_key = quote(f"doi:{DOI}", safe="")
    url = f"{API_ROOT}/datasets/{dataset_key}/download"
    payload = _request_bytes(url, accept="application/zip,application/octet-stream,*/*")
    if not zipfile.is_zipfile(io.BytesIO(payload)):
        raise RuntimeError("Dryad public dataset download did not return a ZIP package")
    return url, payload


def _extract_locked_files(package: bytes) -> dict[str, bytes]:
    expected = {filename for _, filename in FILES}
    found: dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        for info in archive.infolist():
            basename = Path(info.filename).name
            if basename in expected:
                if basename in found:
                    raise RuntimeError(f"duplicate locked filename in Dryad package: {basename}")
                found[basename] = archive.read(info)
    missing = expected - set(found)
    if missing:
        raise RuntimeError(f"locked Miconia files absent from Dryad package: {sorted(missing)}")
    return found


def _schema(path: Path) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    sheets: list[dict[str, object]] = []
    for ws in workbook.worksheets:
        columns: list[str] = []
        if ws.max_row >= 1:
            for cell in ws[1]:
                value = cell.value
                columns.append("" if value is None else str(value))
        sheets.append(
            {
                "sheet": ws.title,
                "rows_including_header": int(ws.max_row),
                "data_rows": max(int(ws.max_row) - 1, 0),
                "columns_count": int(ws.max_column),
                "columns": columns,
            }
        )
    workbook.close()
    return sheets


def discover(output_dir: Path, manifest_path: Path) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)

    dataset_api_url, files_api_url, metadata = _resolve_public_metadata()
    package_url, package = _download_public_package()
    extracted = _extract_locked_files(package)

    metadata_by_name = {str(item["filename"]): item for item in metadata}
    inventory: list[dict[str, object]] = []
    for file_id, filename in FILES:
        payload = extracted[filename]
        target = output_dir / filename
        target.write_bytes(payload)
        meta = metadata_by_name[filename]
        digest = hashlib.sha256(payload).hexdigest()
        metadata_digest = meta.get("metadata_digest")
        metadata_digest_type = str(meta.get("metadata_digest_type") or "").lower()
        if metadata_digest and metadata_digest_type in {"sha-256", "sha256"} and digest != metadata_digest:
            raise RuntimeError(
                f"Dryad package checksum mismatch for {filename}: metadata={metadata_digest}, observed={digest}"
            )
        inventory.append(
            {
                **meta,
                "file_stream_id": file_id,
                "filename": filename,
                "bytes": len(payload),
                "sha256": digest,
                "workbook_schema": _schema(target),
            }
        )

    result: dict[str, object] = {
        "status": "schema_only_discovery_complete",
        "dataset_doi": DOI,
        "dataset_api_url": dataset_api_url,
        "files_api_url": files_api_url,
        "package_download_url": package_url,
        "package_sha256": hashlib.sha256(package).hexdigest(),
        "files": inventory,
        "inspection_boundary": (
            "Only public metadata, workbook names, hashes, dimensions and first-row column labels were inspected. "
            "No data-cell values, outcome summaries, fitted models or effect directions were used."
        ),
        "next_gate": (
            "Map keys and source-defined process variables from schema only, classify joint_state_identifiable / "
            "partial_joint_state_identifiable / not_identifiable_from_archive, then preregister exact models before "
            "any outcome analysis."
        ),
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default="_external/miconia_joint_state")
    parser.add_argument("--manifest", default="artifacts/empirical/miconia_joint_state_schema.json")
    args = parser.parse_args()
    result = discover(Path(args.output_dir), Path(args.manifest))
    print(json.dumps({"status": result["status"], "files": [f["filename"] for f in result["files"]]}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
